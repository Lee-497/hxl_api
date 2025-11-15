"""
库存汇总报表
导入库存数据并剔除指定仓库的数据
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side
from pathlib import Path
from typing import Optional
from config.settings import PROCESSED_DIR
from utils.data_loader import get_data_loader
from utils.data_parser import get_data_parser
from utils.logger import get_logger
from utils.file_utils import generate_timestamped_filename

logger = get_logger(__name__)

# 声明依赖的原始数据模块
DEPENDENCIES = ["inventory_query", "product_archive"]

# 需要剔除的仓库列表
EXCLUDED_WAREHOUSES = [
    "广东从化退货仓",
    "广东东莞二仓退货仓", 
    "东莞中转仓"
]

def run() -> Optional[Path]:
    """
    执行库存汇总报表生成
    
    Returns:
        生成的报表文件路径，失败返回None
    """
    logger.info("开始处理库存报表数据")
    
    try:
        # 1. 加载库存数据
        data_loader = get_data_loader()
        inventory_df = data_loader.load_latest_module_data("inventory_query")
        
        if inventory_df is None or inventory_df.empty:
            logger.error("库存数据加载失败或为空")
            return None
        
        logger.info(f"原始库存数据: {len(inventory_df)} 行")
        
        # 打印数据摘要
        data_parser = get_data_parser()
        data_parser.print_data_summary(inventory_df, "库存数据")
        
        # 2. 加载商品分类数据并关联
        category_df = load_product_categories(data_loader)
        if category_df is not None:
            inventory_df = add_product_categories(inventory_df, category_df)
            logger.info("商品分类关联完成")
        
        # 3. 加载商品人员架构数据并关联采购责任人
        staff_df = load_staff_structure()
        if staff_df is not None:
            inventory_df = add_purchase_manager(inventory_df, staff_df)
            logger.info("采购责任人关联完成")
        
        # 4. 数据清洗：剔除指定仓库的数据
        filtered_df = filter_excluded_warehouses(inventory_df)
        
        # 4.5. 数据清洗：剔除指定一级分类的数据
        filtered_df = filter_excluded_categories(filtered_df)
        
        # 5. 加载门店商品属性数据并关联
        attr_df = load_store_product_attributes(data_loader)
        if attr_df is not None:
            filtered_df = add_store_product_attributes(filtered_df, attr_df)
            logger.info("门店商品属性关联完成")
        
        # 6. 数据转换：将门店从行转为列
        pivoted_df = pivot_stores_to_columns(filtered_df, attr_df)
        logger.info(f"透视转换后数据: {len(pivoted_df)} 行, {len(pivoted_df.columns)} 列")
        
        # 6. 保存处理后的报表
        output_filename = generate_timestamped_filename("库存汇总报表", "xlsx")
        output_path = PROCESSED_DIR / output_filename
        
        # 在保存前最后一次处理空值，确保Excel中显示为空白而不是NaN
        final_df = pivoted_df.copy()
        special_columns = ['一级分类', '二级分类', '采购责任人']
        for col in special_columns:
            if col in final_df.columns:
                final_df[col] = final_df[col].fillna('').astype(str).replace('nan', '')
        
        # 处理门店属性列
        store_attr_columns = [col for col in final_df.columns if '_停购' in col or '_停止要货' in col]
        for col in store_attr_columns:
            if col in final_df.columns:
                final_df[col] = final_df[col].fillna('').astype(str).replace('nan', '')
        
        final_df.to_excel(output_path, index=False, engine='openpyxl')
        apply_inventory_report_style(output_path)
        
        logger.info(f"库存汇总报表生成完成: {output_path}")
        print(f"[完成] 库存汇总报表: {output_filename}")
        
        return output_path
        
    except Exception as e:
        logger.error(f"生成库存汇总报表时发生异常: {str(e)}")
        return None


def apply_inventory_report_style(file_path: Path) -> None:
    try:
        workbook = load_workbook(file_path)
        thin_side = Side(style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        font = Font(name="微软雅黑", size=10)

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.border = border

        workbook.save(file_path)
        workbook.close()
    except Exception as exc:
        logger.warning(f"应用库存报表样式失败: {exc}")


def filter_excluded_warehouses(inventory_df: pd.DataFrame) -> pd.DataFrame:
    """
    过滤掉指定的仓库数据
    
    Args:
        inventory_df: 库存数据DataFrame
        
    Returns:
        过滤后的DataFrame
    """
    if inventory_df.empty:
        logger.warning("库存数据为空，跳过仓库过滤")
        return inventory_df
    
    # 尝试找到仓库字段（可能的字段名）
    warehouse_columns = ['仓库', '仓库名称', '仓库名', 'warehouse', 'warehouse_name', '仓库编码']
    data_parser = get_data_parser()
    warehouse_col = data_parser.find_column(inventory_df, warehouse_columns)
    
    if warehouse_col is None:
        logger.warning(f"未找到仓库字段，可用字段: {list(inventory_df.columns)}")
        return inventory_df
    
    # 记录过滤前的数据量
    original_count = len(inventory_df)
    
    # 过滤数据
    filtered_df = inventory_df[~inventory_df[warehouse_col].isin(EXCLUDED_WAREHOUSES)]
    
    # 记录过滤结果
    filtered_count = original_count - len(filtered_df)
    if filtered_count > 0:
        logger.info(f"已剔除 {filtered_count} 条数据（仓库: {', '.join(EXCLUDED_WAREHOUSES)}）")
        print(f"[过滤] 剔除指定仓库数据 {filtered_count} 条")
        
        # 显示被剔除的仓库统计
        excluded_data = inventory_df[inventory_df[warehouse_col].isin(EXCLUDED_WAREHOUSES)]
        warehouse_counts = excluded_data[warehouse_col].value_counts()
        for warehouse, count in warehouse_counts.items():
            print(f"   - {warehouse}: {count} 条")
    else:
        logger.info("未发现需要剔除的仓库数据")
    
    return filtered_df


def filter_excluded_categories(inventory_df: pd.DataFrame) -> pd.DataFrame:
    """
    过滤掉指定一级分类的数据
    
    Args:
        inventory_df: 库存数据DataFrame
        
    Returns:
        过滤后的DataFrame
    """
    if inventory_df.empty:
        logger.warning("库存数据为空，跳过分类过滤")
        return inventory_df
    
    # 定义需要剔除的一级分类
    EXCLUDED_CATEGORIES = ['冷藏食品', '冷冻食品']
    
    # 查找一级分类字段
    category_col = None
    if '一级分类' in inventory_df.columns:
        category_col = '一级分类'
    else:
        logger.warning("未找到一级分类字段，跳过分类过滤")
        return inventory_df
    
    # 记录过滤前的数据量
    original_count = len(inventory_df)
    
    # 过滤数据：剔除指定的一级分类
    filtered_df = inventory_df[~inventory_df[category_col].isin(EXCLUDED_CATEGORIES)]
    
    # 记录过滤结果
    filtered_count = original_count - len(filtered_df)
    if filtered_count > 0:
        logger.info(f"已剔除 {filtered_count} 条数据（一级分类: {', '.join(EXCLUDED_CATEGORIES)}）")
        print(f"[过滤] 剔除指定分类数据 {filtered_count} 条")
        
        # 显示被剔除的分类统计
        excluded_data = inventory_df[inventory_df[category_col].isin(EXCLUDED_CATEGORIES)]
        category_counts = excluded_data[category_col].value_counts()
        for category, count in category_counts.items():
            print(f"   - {category}: {count} 条")
    else:
        logger.info("未发现需要剔除的分类数据")
    
    return filtered_df


def load_product_categories(data_loader) -> Optional[pd.DataFrame]:
    """
    加载商品分类数据
    
    Args:
        data_loader: 数据加载器实例
        
    Returns:
        商品分类DataFrame，失败返回None
    """
    try:
        # 尝试从组织商品档案加载分类数据
        category_df = data_loader.load_latest_module_data("product_archive")
        
        if category_df is None or category_df.empty:
            logger.error("无法加载商品分类数据")
            return None
        
        # 选择需要的字段
        required_columns = ['商品代码', '一级分类', '二级分类']
        missing_columns = [col for col in required_columns if col not in category_df.columns]
        
        if missing_columns:
            logger.error(f"商品分类数据缺少必需字段: {missing_columns}")
            return None
        
        # 筛选有效的分类数据
        category_clean = category_df[required_columns].dropna()
        category_clean = category_clean[category_clean['商品代码'] != '']
        
        logger.info(f"成功加载商品分类数据: {len(category_clean)} 条")
        return category_clean
        
    except Exception as e:
        logger.error(f"加载商品分类数据时发生异常: {str(e)}")
        return None


def add_product_categories(inventory_df: pd.DataFrame, category_df: pd.DataFrame) -> pd.DataFrame:
    """
    为库存数据添加商品分类信息
    
    Args:
        inventory_df: 库存数据DataFrame
        category_df: 商品分类数据DataFrame
        
    Returns:
        添加了分类信息的DataFrame
    """
    if inventory_df.empty or category_df.empty:
        logger.warning("库存数据或分类数据为空，跳过分类关联")
        return inventory_df
    
    try:
        # 确保商品代码字段存在
        if '商品代码' not in inventory_df.columns:
            logger.error("库存数据中缺少商品代码字段")
            return inventory_df
        
        # 执行左连接，保留所有库存数据
        merged_df = inventory_df.merge(
            category_df[['商品代码', '一级分类', '二级分类']], 
            on='商品代码', 
            how='left'
        )
        
        # 立即处理合并后的空值，避免后续转换为字符串'nan'
        merged_df['一级分类'] = merged_df['一级分类'].fillna('')
        merged_df['二级分类'] = merged_df['二级分类'].fillna('')
        
        # 统计关联结果
        total_count = len(inventory_df)
        matched_count = (merged_df['一级分类'] != '').sum()
        match_rate = (matched_count / total_count * 100) if total_count > 0 else 0
        
        logger.info(f"分类关联完成: {matched_count}/{total_count} ({match_rate:.1f}%) 商品匹配到分类")
        print(f"[关联] 商品分类匹配: {matched_count}/{total_count} ({match_rate:.1f}%)")
        
        return merged_df
        
    except Exception as e:
        logger.error(f"关联商品分类时发生异常: {str(e)}")
        return inventory_df


def load_staff_structure() -> Optional[pd.DataFrame]:
    """
    加载商品人员架构表数据
    
    Returns:
        商品人员架构DataFrame，失败返回None
    """
    try:
        # 使用相对路径查找商品人员架构文件
        from pathlib import Path
        base_dir = Path(__file__).parent.parent
        staff_file = base_dir / "storage" / "reference" / "商品人员架构.xlsx"
        
        if not staff_file.exists():
            logger.warning(f"商品人员架构文件不存在: {staff_file}")
            return None
            
        staff_df = pd.read_excel(staff_file)
        
        if staff_df.empty:
            logger.error("商品人员架构表数据为空")
            return None
        
        # 选择需要的字段
        required_columns = ['二级分类', '采购责任人']
        missing_columns = [col for col in required_columns if col not in staff_df.columns]
        
        if missing_columns:
            logger.error(f"商品人员架构表缺少必需字段: {missing_columns}")
            return None
        
        # 筛选有效的人员数据，去重
        staff_clean = staff_df[required_columns].dropna()
        staff_clean = staff_clean[staff_clean['二级分类'] != '']
        staff_clean = staff_clean.drop_duplicates(subset=['二级分类'])
        
        logger.info(f"成功加载商品人员架构数据: {len(staff_clean)} 条")
        return staff_clean
        
    except Exception as e:
        logger.error(f"加载商品人员架构数据时发生异常: {str(e)}")
        return None


def add_purchase_manager(inventory_df: pd.DataFrame, staff_df: pd.DataFrame) -> pd.DataFrame:
    """
    为库存数据添加采购责任人信息
    
    Args:
        inventory_df: 库存数据DataFrame
        staff_df: 商品人员架构数据DataFrame
        
    Returns:
        添加了采购责任人信息的DataFrame
    """
    if inventory_df.empty or staff_df.empty:
        logger.warning("库存数据或人员架构数据为空，跳过采购责任人关联")
        return inventory_df
    
    try:
        # 确保二级分类字段存在
        if '二级分类' not in inventory_df.columns:
            logger.error("库存数据中缺少二级分类字段")
            return inventory_df
        
        # 执行左连接，保留所有库存数据
        merged_df = inventory_df.merge(
            staff_df[['二级分类', '采购责任人']], 
            on='二级分类', 
            how='left'
        )
        
        # 立即处理合并后的空值
        merged_df['采购责任人'] = merged_df['采购责任人'].fillna('')
        
        # 统计关联结果
        total_count = len(inventory_df)
        matched_count = (merged_df['采购责任人'] != '').sum()
        match_rate = (matched_count / total_count * 100) if total_count > 0 else 0
        
        logger.info(f"采购责任人关联完成: {matched_count}/{total_count} ({match_rate:.1f}%) 商品匹配到责任人")
        print(f"[关联] 采购责任人匹配: {matched_count}/{total_count} ({match_rate:.1f}%)")
        
        return merged_df
        
    except Exception as e:
        logger.error(f"关联采购责任人时发生异常: {str(e)}")
        return inventory_df


def load_store_product_attributes(data_loader) -> Optional[pd.DataFrame]:
    """
    加载门店商品属性表数据
    
    Args:
        data_loader: 数据加载器实例
        
    Returns:
        门店商品属性DataFrame，失败返回None
    """
    try:
        # 尝试从模块加载门店商品属性数据
        attr_df = data_loader.load_latest_module_data("store_product_attributes")
        
        if attr_df is None or attr_df.empty:
            logger.error("无法加载门店商品属性数据")
            return None
        
        # 选择需要的字段
        required_columns = ['门店', '商品代码', '停购', '停止要货']
        missing_columns = [col for col in required_columns if col not in attr_df.columns]
        
        if missing_columns:
            logger.error(f"门店商品属性数据缺少必需字段: {missing_columns}")
            return None
        
        # 筛选有效的属性数据
        # 注意：不能使用dropna()，因为这会删除停购/停止要货为空值的记录
        # 只过滤掉门店和商品代码为空的记录
        attr_clean = attr_df[required_columns].copy()
        attr_clean = attr_clean[
            (attr_clean['门店'].notna()) & 
            (attr_clean['门店'] != '') & 
            (attr_clean['商品代码'].notna()) & 
            (attr_clean['商品代码'] != '')
        ]
        
        logger.info(f"成功加载门店商品属性数据: {len(attr_clean)} 条")
        
        # 调试：检查属性数据的分布情况
        if not attr_clean.empty:
            store_count = attr_clean['门店'].nunique()
            product_count = attr_clean['商品代码'].nunique()
            total_combinations = store_count * product_count
            actual_records = len(attr_clean)
            
            logger.info(f"门店商品属性数据覆盖: {store_count} 个门店, {product_count} 个商品")
            logger.info(f"理论上应有 {total_combinations} 条记录，实际有 {actual_records} 条记录")
            
            if actual_records < total_combinations:
                logger.warning(f"门店商品属性表数据不完整！缺少 {total_combinations - actual_records} 条记录")
                
                # 分析哪些门店+商品组合缺失
                all_stores = attr_clean['门店'].unique()
                all_products = attr_clean['商品代码'].unique()
                
                # 检查每个门店的商品覆盖情况
                for store in all_stores[:3]:  # 只检查前3个门店作为样本
                    store_products = attr_clean[attr_clean['门店'] == store]['商品代码'].nunique()
                    missing_products = product_count - store_products
                    if missing_products > 0:
                        logger.warning(f"门店 '{store}' 缺少 {missing_products} 个商品的属性记录")
                
                # 检查每个商品的门店覆盖情况
                sample_products = all_products[:5]  # 检查前5个商品作为样本
                for product in sample_products:
                    product_stores = attr_clean[attr_clean['商品代码'] == product]['门店'].unique()
                    missing_stores = set(all_stores) - set(product_stores)
                    if missing_stores:
                        logger.warning(f"商品 '{product}' 在门店 {list(missing_stores)} 中缺少属性记录")
            
            # 检查停购和停止要货字段的空值情况
            stop_purchase_null = attr_clean['停购'].isna().sum()
            stop_order_null = attr_clean['停止要货'].isna().sum()
            logger.info(f"属性字段空值统计: 停购={stop_purchase_null}, 停止要货={stop_order_null}")
            
            # 检查停购和停止要货字段的值分布
            if stop_purchase_null < len(attr_clean):
                stop_purchase_values = attr_clean['停购'].value_counts(dropna=False)
                logger.info(f"停购字段值分布: {dict(stop_purchase_values)}")
            
            if stop_order_null < len(attr_clean):
                stop_order_values = attr_clean['停止要货'].value_counts(dropna=False)
                logger.info(f"停止要货字段值分布: {dict(stop_order_values)}")
        
        return attr_clean
        
    except Exception as e:
        logger.error(f"加载门店商品属性数据时发生异常: {str(e)}")
        return None


def add_store_product_attributes(inventory_df: pd.DataFrame, attr_df: pd.DataFrame) -> pd.DataFrame:
    """
    为库存数据添加门店商品属性信息
    
    Args:
        inventory_df: 库存数据DataFrame
        attr_df: 门店商品属性数据DataFrame
        
    Returns:
        添加了门店商品属性的DataFrame
    """
    if inventory_df.empty or attr_df.empty:
        logger.warning("库存数据或门店商品属性数据为空，跳过门店属性关联")
        return inventory_df
    
    try:
        # 确保必需字段存在
        required_fields = ['门店', '商品代码']
        missing_fields = [field for field in required_fields if field not in inventory_df.columns]
        
        if missing_fields:
            logger.error(f"库存数据中缺少必需字段: {missing_fields}")
            return inventory_df
        
        # 关键修复：先清除库存表中可能存在的停购、停止要货字段，确保属性表是唯一数据源
        inventory_clean = inventory_df.copy()
        if '停购' in inventory_clean.columns:
            inventory_clean = inventory_clean.drop(columns=['停购'])
            logger.info("已清除库存表中的停购字段，将使用属性表作为唯一数据源")
        if '停止要货' in inventory_clean.columns:
            inventory_clean = inventory_clean.drop(columns=['停止要货'])
            logger.info("已清除库存表中的停止要货字段，将使用属性表作为唯一数据源")
        
        # 执行左连接，同时匹配门店和商品代码
        logger.info(f"开始关联门店商品属性，库存数据: {len(inventory_clean)} 条，属性数据: {len(attr_df)} 条")
        
        merged_df = inventory_clean.merge(
            attr_df[['门店', '商品代码', '停购', '停止要货']], 
            on=['门店', '商品代码'], 
            how='left'
        )
        
        logger.info(f"关联后数据: {len(merged_df)} 条")
        
        # 调试：检查关联结果，特别关注同一商品在不同门店的匹配情况
        unmatched_before = merged_df['停购'].isna().sum()
        logger.info(f"关联后未匹配的记录数: {unmatched_before}")
        
        # 详细分析库存表和属性表的匹配情况
        inventory_stores = set(inventory_df['门店'].unique())
        inventory_products = set(inventory_df['商品代码'].unique())
        attr_stores = set(attr_df['门店'].unique())
        attr_products = set(attr_df['商品代码'].unique())
        
        logger.info(f"库存表覆盖: {len(inventory_stores)} 个门店, {len(inventory_products)} 个商品")
        logger.info(f"属性表覆盖: {len(attr_stores)} 个门店, {len(attr_products)} 个商品")
        
        # 检查门店和商品的交集
        common_stores = inventory_stores & attr_stores
        common_products = inventory_products & attr_products
        logger.info(f"共同门店: {len(common_stores)} 个, 共同商品: {len(common_products)} 个")
        
        if len(common_stores) < len(inventory_stores):
            missing_stores = inventory_stores - attr_stores
            logger.warning(f"属性表中缺少的门店: {list(missing_stores)}")
        
        if len(common_products) < len(inventory_products):
            missing_products = inventory_products - attr_products
            logger.warning(f"属性表中缺少的商品数量: {len(missing_products)}")
        
        # 随机选择几个商品检查其在不同门店的匹配情况
        sample_products = merged_df['商品代码'].unique()[:3]
        for product_code in sample_products:
            product_records = merged_df[merged_df['商品代码'] == product_code]
            matched_stores = product_records[product_records['停购'].notna()]
            unmatched_stores = product_records[product_records['停购'].isna()]
            
            if len(matched_stores) > 0 and len(unmatched_stores) > 0:
                logger.warning(f"商品 {product_code}: {len(matched_stores)} 个门店有属性, {len(unmatched_stores)} 个门店无属性")
                logger.warning(f"  有属性的门店: {list(matched_stores['门店'].unique())}")
                logger.warning(f"  无属性的门店: {list(unmatched_stores['门店'].unique())}")
                
                # 检查这个商品在属性表中的实际情况
                product_in_attr = attr_df[attr_df['商品代码'] == product_code]
                if not product_in_attr.empty:
                    attr_stores_for_product = set(product_in_attr['门店'].unique())
                    inventory_stores_for_product = set(product_records['门店'].unique())
                    missing_attr_stores = inventory_stores_for_product - attr_stores_for_product
                    if missing_attr_stores:
                        logger.warning(f"  商品 {product_code} 在属性表中缺少门店: {list(missing_attr_stores)}")
                else:
                    logger.warning(f"  商品 {product_code} 在属性表中完全不存在！")
        
        # 处理合并后的空值：只有完全匹配不上的才设为空，匹配上但值为null的保持原样
        # 注意：merge后，匹配不上的记录这两个字段会是NaN，匹配上但原值为空的会保持原来的空值
        merged_df['停购'] = merged_df['停购'].fillna('')
        merged_df['停止要货'] = merged_df['停止要货'].fillna('')
        
        # 统计关联结果
        total_count = len(inventory_df)
        matched_count = (merged_df['停购'] != '').sum()
        match_rate = (matched_count / total_count * 100) if total_count > 0 else 0
        
        logger.info(f"门店商品属性关联完成: {matched_count}/{total_count} ({match_rate:.1f}%) 记录匹配到属性")
        print(f"[关联] 门店商品属性匹配: {matched_count}/{total_count} ({match_rate:.1f}%)")
        
        return merged_df
        
    except Exception as e:
        logger.error(f"关联门店商品属性时发生异常: {str(e)}")
        return inventory_df


def pivot_stores_to_columns(inventory_df: pd.DataFrame, attr_df: pd.DataFrame = None) -> pd.DataFrame:
    """
    将门店字段从行转换为列，实现数据透视
    
    Args:
        inventory_df: 过滤后的库存数据DataFrame
        attr_df: 门店商品属性数据DataFrame，用于修复透视表中的空值
        
    Returns:
        透视后的DataFrame，门店作为列，商品信息作为行
    """
    if inventory_df.empty:
        logger.warning("库存数据为空，跳过透视转换")
        return inventory_df
    
    # 定义需要保留的字段
    required_fields = ['商品代码', '商品条码', '商品名称', '数量', '可用数量']
    category_fields = ['一级分类', '二级分类']  # 分类字段
    staff_fields = ['采购责任人']  # 人员字段
    attribute_fields = ['停购', '停止要货']  # 门店商品属性字段
    store_field = '门店'
    
    # 检查必需字段是否存在
    missing_fields = []
    for field in required_fields + [store_field]:
        if field not in inventory_df.columns:
            missing_fields.append(field)
    
    if missing_fields:
        logger.error(f"缺少必需字段: {missing_fields}")
        logger.info(f"可用字段: {list(inventory_df.columns)}")
        return inventory_df
    
    try:
        # 确定实际可用的字段
        available_fields = required_fields + [store_field]
        available_category_fields = []
        available_staff_fields = []
        available_attribute_fields = []
        
        # 添加分类字段
        for field in category_fields:
            if field in inventory_df.columns:
                available_fields.append(field)
                available_category_fields.append(field)
        
        # 添加人员字段
        for field in staff_fields:
            if field in inventory_df.columns:
                available_fields.append(field)
                available_staff_fields.append(field)
        
        # 添加门店商品属性字段
        for field in attribute_fields:
            if field in inventory_df.columns:
                available_fields.append(field)
                available_attribute_fields.append(field)
        
        # 选择需要的字段
        selected_df = inventory_df[available_fields].copy()
        
        # 调试：检查字段识别情况
        logger.info(f"可用字段: {available_fields}")
        logger.info(f"分类字段: {available_category_fields}")
        logger.info(f"人员字段: {available_staff_fields}")
        logger.info(f"属性字段: {available_attribute_fields}")
        
        # 检查属性字段的数据情况
        for field in available_attribute_fields:
            field_values = selected_df[field].value_counts(dropna=False)
            logger.info(f"字段 {field} 值分布: {dict(field_values.head(5))}")
        
        # 处理分类字段、人员字段和属性字段的空值，将NaN替换为空字符串，避免groupby时被排除
        for field in available_category_fields + available_staff_fields + available_attribute_fields:
            selected_df[field] = selected_df[field].fillna('').astype(str)
            # 处理字符串形式的'nan'
            selected_df[field] = selected_df[field].replace('nan', '')
        
        # 按商品信息分组，对数量和可用数量分别求和（处理同一商品在同一门店的多条记录）
        # 注意：不在group_fields中包含属性字段，因为它们要按门店展开
        group_fields = ['商品代码', '商品条码', '商品名称'] + available_category_fields + available_staff_fields + ['门店']
        
        # 对数量字段求和，对属性字段取第一个值（因为同一商品在同一门店的属性应该一致）
        agg_dict = {'数量': 'sum', '可用数量': 'sum'}
        for field in available_attribute_fields:
            agg_dict[field] = 'first'
        
        grouped_df = selected_df.groupby(group_fields, dropna=False).agg(agg_dict).reset_index()
        
        # 执行透视转换：门店作为列，商品信息、分类、人员作为行索引（不包含属性字段，因为它们要按门店展开）
        index_fields = ['商品代码', '商品条码', '商品名称'] + available_category_fields + available_staff_fields
        
        # 分别对数量、可用数量、停购、停止要货进行透视
        pivoted_qty = grouped_df.pivot_table(
            index=index_fields,
            columns='门店',
            values='数量',
            fill_value=0,
            aggfunc='sum'
        )
        
        pivoted_available = grouped_df.pivot_table(
            index=index_fields,
            columns='门店',
            values='可用数量',
            fill_value=0,
            aggfunc='sum'
        )
        
        # 调试：检查分组后的停购/停止要货数据
        logger.info(f"分组后数据: {len(grouped_df)} 行")
        stop_purchase_values = grouped_df['停购'].value_counts(dropna=False)
        stop_order_values = grouped_df['停止要货'].value_counts(dropna=False)
        logger.info(f"分组后停购值分布: {dict(stop_purchase_values)}")
        logger.info(f"分组后停止要货值分布: {dict(stop_order_values)}")
        
        # 对停购和停止要货进行特殊处理：逐商品查找属性表，确保数据完整性
        # 先创建基础的透视表结构
        pivoted_stop_purchase = grouped_df.pivot_table(
            index=index_fields,
            columns='门店',
            values='停购',
            fill_value='',
            aggfunc='first'
        )
        
        pivoted_stop_order = grouped_df.pivot_table(
            index=index_fields,
            columns='门店',
            values='停止要货',
            fill_value='',
            aggfunc='first'
        )
        
        # 关键修复：对于空值，直接从原始属性表中查找
        if attr_df is not None and not attr_df.empty:
            logger.info("开始修复透视表中的空值，直接从属性表查找...")
            
            for store in pivoted_stop_purchase.columns:
                stop_purchase_col = f'{store}_停购'
                stop_order_col = f'{store}_停止要货'
                
                # 统计修复前的空值数量
                empty_purchase_before = (pivoted_stop_purchase[store] == '').sum()
                empty_order_before = (pivoted_stop_order[store] == '').sum()
                
                if empty_purchase_before > 0 or empty_order_before > 0:
                    logger.info(f"门店 {store}: 停购空值 {empty_purchase_before} 个, 停止要货空值 {empty_order_before} 个")
                
                    # 对每个空值记录，直接从属性表查找
                    for idx in pivoted_stop_purchase.index:
                        # 获取商品代码（索引的第一个元素）
                        product_code = idx[0] if isinstance(idx, tuple) else idx
                        
                        # 如果当前值为空，从属性表查找
                        if pivoted_stop_purchase.loc[idx, store] == '':
                            matched_data = attr_df[
                                (attr_df['商品代码'] == product_code) &
                                (attr_df['门店'] == store)
                            ]
                            if not matched_data.empty:
                                stop_purchase_val = matched_data['停购'].iloc[0]
                                if pd.notna(stop_purchase_val) and stop_purchase_val != '':
                                    pivoted_stop_purchase.loc[idx, store] = stop_purchase_val
                        
                        if pivoted_stop_order.loc[idx, store] == '':
                            matched_data = attr_df[
                                (attr_df['商品代码'] == product_code) &
                                (attr_df['门店'] == store)
                            ]
                            if not matched_data.empty:
                                stop_order_val = matched_data['停止要货'].iloc[0]
                                if pd.notna(stop_order_val) and stop_order_val != '':
                                    pivoted_stop_order.loc[idx, store] = stop_order_val
                    
                    # 统计修复后的空值数量
                    empty_purchase_after = (pivoted_stop_purchase[store] == '').sum()
                    empty_order_after = (pivoted_stop_order[store] == '').sum()
                    
                    fixed_purchase = empty_purchase_before - empty_purchase_after
                    fixed_order = empty_order_before - empty_order_after
                    
                    if fixed_purchase > 0 or fixed_order > 0:
                        logger.info(f"门店 {store}: 修复了 停购 {fixed_purchase} 个, 停止要货 {fixed_order} 个空值")
        else:
            logger.warning("属性表数据不可用，跳过空值修复")
        
        # 调试：检查透视后的空值情况
        for col in pivoted_stop_purchase.columns:
            empty_count = (pivoted_stop_purchase[col] == '').sum()
            total_count = len(pivoted_stop_purchase)
            if empty_count > 0:
                logger.warning(f"透视后 {col}_停购 有 {empty_count}/{total_count} 个空值")
        
        for col in pivoted_stop_order.columns:
            empty_count = (pivoted_stop_order[col] == '').sum()
            total_count = len(pivoted_stop_order)
            if empty_count > 0:
                logger.warning(f"透视后 {col}_停止要货 有 {empty_count}/{total_count} 个空值")
        
        # 重命名列名，区分不同类型的字段
        pivoted_qty.columns = [f'{col}_数量' for col in pivoted_qty.columns]
        pivoted_available.columns = [f'{col}_可用数量' for col in pivoted_available.columns]
        pivoted_stop_purchase.columns = [f'{col}_停购' for col in pivoted_stop_purchase.columns]
        pivoted_stop_order.columns = [f'{col}_停止要货' for col in pivoted_stop_order.columns]
        
        # 合并所有透视表
        pivoted_df = pivoted_qty.join(pivoted_available).join(pivoted_stop_purchase).join(pivoted_stop_order)
        
        # 重置索引，将商品信息从索引转为普通列
        pivoted_df = pivoted_df.reset_index()
        
        # 重命名列索引名称（去掉'门店'这个列名）
        pivoted_df.columns.name = None
        
        # 重新排列列的顺序，让每个门店的所有字段相邻
        base_columns = ['商品代码', '商品条码', '商品名称'] + available_category_fields + available_staff_fields
        store_names = list(set([col.split('_')[0] for col in pivoted_df.columns if '_数量' in col or '_可用数量' in col or '_停购' in col or '_停止要货' in col]))
        store_names.sort()  # 按门店名称排序
        
        # 构建新的列顺序
        new_columns = base_columns.copy()
        for store in store_names:
            qty_col = f'{store}_数量'
            available_col = f'{store}_可用数量'
            stop_purchase_col = f'{store}_停购'
            stop_order_col = f'{store}_停止要货'
            
            if qty_col in pivoted_df.columns:
                new_columns.append(qty_col)
            if available_col in pivoted_df.columns:
                new_columns.append(available_col)
            if stop_purchase_col in pivoted_df.columns:
                new_columns.append(stop_purchase_col)
            if stop_order_col in pivoted_df.columns:
                new_columns.append(stop_order_col)
        
        # 重新排列DataFrame的列
        pivoted_df = pivoted_df[new_columns]
        
        # 保留库存为0的记录，不转换为空值（库存为0是有意义的信息）
        # 注释掉原来的逻辑，保持数量字段的原始值
        # store_columns = [col for col in pivoted_df.columns if '_数量' in col or '_可用数量' in col]
        # for col in store_columns:
        #     pivoted_df[col] = pivoted_df[col].replace(0, '')
        
        # 处理分类字段和人员字段中的'nan'字符串，转换为空值
        for field in available_category_fields + available_staff_fields:
            if field in pivoted_df.columns:
                pivoted_df[field] = pivoted_df[field].replace('nan', '')
        
        # 处理门店属性字段：只处理字符串形式的'nan'，保留其他值
        store_attr_columns = [col for col in pivoted_df.columns if '_停购' in col or '_停止要货' in col]
        for col in store_attr_columns:
            # 统计最终的值分布
            final_values = pivoted_df[col].value_counts(dropna=False)
            logger.info(f"列 {col} 最终值分布: {dict(final_values.head(10))}")
            
            # 只处理字符串形式的'nan'，不处理其他空值
            pivoted_df[col] = pivoted_df[col].replace('nan', '')
        
        # 获取所有门店相关的列用于日志输出
        all_store_columns = [col for col in pivoted_df.columns if any(store in col for store in store_names)]
        logger.info(f"透视转换完成，门店相关列数: {len(all_store_columns)}")
        print(f"[转换] 门店字段已展开为列，共 {len(all_store_columns)} 个门店相关字段")
        
        return pivoted_df
        
    except Exception as e:
        logger.error(f"透视转换失败: {str(e)}")
        return inventory_df



def get_dependencies() -> list:
    """获取依赖列表"""
    return DEPENDENCIES


def get_description() -> str:
    """获取报表描述"""
    return "导入库存数据，关联商品分类，剔除指定仓库，并将门店字段展开为列的透视报表"