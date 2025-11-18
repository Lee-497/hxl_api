"""冷藏乳饮销售分析报表加工"""

from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

from config.settings import PROCESSED_DIR
from utils.data_loader import get_data_loader
from utils.logger import get_logger

logger = get_logger(__name__)

# 声明依赖的原始数据模块
DEPENDENCIES = ["商品销售数据"]

TARGET_COLUMNS: List[str] = [
    "门店代码",
    "门店名称",
    "一级类别",
    "二级类别",
    "三级类别",
    "商品代码",
    "商品条码",
    "商品名称",
    "采购规格",
    "基本单位",
    "数量合计",
    "金额合计",
]

EXCLUDE_KEYWORDS: List[str] = ["益力多"]


def run() -> Optional[Path]:
    """
    执行销售分析报表生成
    
    Returns:
        生成的报表文件路径，失败返回None
    """
    logger.info("开始加工冷藏乳饮销售分析报表")

    try:
        # 1. 加载销售分析数据
        data_loader = get_data_loader()
        sales_df = data_loader.load_latest_module_data("商品销售数据")

        if sales_df is None or sales_df.empty:
            logger.error("销售分析数据加载失败或为空")
            return None

        # 2. 数据过滤与字段保留
        processed_df = transform_sales_data(sales_df)
        if processed_df is None or processed_df.empty:
            logger.error("冷藏乳饮报表处理后无有效数据")
            return None

        logger.info(f"处理后数据量: {len(processed_df)} 行")

        # 3. 保存处理后的报表
        report_date = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        base_filename = f"冷藏乳饮{report_date}.xlsx"
        output_path = PROCESSED_DIR / base_filename

        # 确保目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # 避免同日重复生成时覆盖
        suffix = 1
        while output_path.exists():
            output_path = PROCESSED_DIR / f"冷藏乳饮{report_date}_{suffix}.xlsx"
            suffix += 1
        
        processed_df.to_excel(output_path, sheet_name='冷藏乳饮销售明细', index=False)

        apply_sheet_style(output_path, '冷藏乳饮销售明细')

        logger.info(f"冷藏乳饮销售分析报表生成成功: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"冷藏乳饮销售分析报表生成失败: {str(e)}")
        return None


def transform_sales_data(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    数据过滤与字段保留
    
    Args:
        df: 原始销售数据
        
    Returns:
        处理后的数据框
    """
    try:
        processed_df = df.copy()

        # 过滤益力多相关商品
        if "商品名称" in processed_df.columns and EXCLUDE_KEYWORDS:
            pattern = "|".join(EXCLUDE_KEYWORDS)
            mask = processed_df["商品名称"].astype(str).str.contains(pattern, case=False, na=False)
            filtered_count = mask.sum()
            if filtered_count:
                logger.info(f"过滤益力多相关记录 {filtered_count} 条")
            processed_df = processed_df[~mask]
        else:
            logger.warning("未找到'商品名称'列，无法过滤益力多相关商品")

        # 校验必需字段
        available_columns = [col for col in TARGET_COLUMNS if col in processed_df.columns]
        missing_columns = [col for col in TARGET_COLUMNS if col not in processed_df.columns]
        if missing_columns:
            logger.warning(f"缺少以下列，结果中将无法包含: {', '.join(missing_columns)}")

        if not available_columns:
            logger.error("目标字段全部缺失，无法生成报表")
            return None

        processed_df = processed_df[available_columns]

        return processed_df.reset_index(drop=True)
        
    except Exception as e:
        logger.error(f"数据过滤与字段保留失败: {str(e)}")
        return None


def apply_sheet_style(file_path: Path, sheet_name: str) -> None:
    """为生成的报表应用边框与字体样式"""
    try:
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"工作表 {sheet_name} 不存在，跳过样式应用")
            return

        sheet = workbook[sheet_name]
        font = Font(name="微软雅黑", size=10)
        thin_side = Side(style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.border = border

        workbook.save(file_path)
        workbook.close()
    except Exception as exc:
        logger.warning(f"应用表格样式失败: {exc}")


def get_description() -> str:
    """获取报表描述"""
    return "冷藏乳饮销售分析数据加工（剔除益力多相关商品）"
