"""库存门店分类透视报表
根据库存数据生成按仓库-门店-一级分类的可用数量与金额透视表"""

from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side

from config.settings import PROCESSED_DIR
from utils.data_loader import get_data_loader
from utils.data_parser import get_data_parser
from utils.file_utils import generate_timestamped_filename
from utils.logger import get_logger

logger = get_logger(__name__)

# 原始数据依赖
DEPENDENCIES = ["inventory_query", "product_archive", "delivery_analysis"]

# 需要剔除的仓库
EXCLUDED_WAREHOUSES = [
    "广东从化退货仓",
    "广东东莞二仓退货仓",
    "东莞中转仓",
]

WAREHOUSE_CANDIDATES = ["仓库", "仓库名称", "仓库名", "warehouse", "warehouse_name", "仓库编码"]
STORE_CANDIDATES = ["门店", "门店名称", "store", "store_name", "门店编码"]
CATEGORY_CANDIDATES = ["一级分类", "一级类别", "一级品类", "category_lv1", "category_level1"]
PRODUCT_CODE_CANDIDATES = ["商品代码", "item_code", "商品编号"]
AVAILABLE_QTY_CANDIDATES = ["可用数量", "数量", "available_qty", "可售数量"]
AVAILABLE_AMOUNT_CANDIDATES = ["可用金额", "金额", "available_amount", "可售金额"]

DELIVERY_STORE_CANDIDATES = ["调出门店", "调出门店名称", "门店", "出货门店"]
DELIVERY_CATEGORY_CANDIDATES = [
    "商品类别名称",
    "一级分类",
    "一级类别",
    "一级品类",
    "商品类别",
]
DELIVERY_QTY_CANDIDATES = ["配送数量", "配送合计数量", "数量"]
DELIVERY_AMOUNT_CANDIDATES = ["配送金额", "配送合计金额", "金额"]


def run() -> Optional[Path]:
    """生成库存门店分类透视报表"""
    logger.info("开始生成库存门店分类透视报表")

    try:
        data_loader = get_data_loader()
        data_parser = get_data_parser()

        inventory_df = data_loader.load_latest_module_data("inventory_query")
        if inventory_df is None or inventory_df.empty:
            logger.error("库存数据加载失败或为空")
            return None

        logger.info(f"库存数据量: {len(inventory_df)} 行")

        # 加载一级分类数据
        category_df = load_product_categories(data_loader)
        if category_df is not None:
            inventory_df = add_product_categories(inventory_df, category_df)

        # 定位关键字段
        warehouse_col = data_parser.find_column(inventory_df, WAREHOUSE_CANDIDATES)
        store_col = data_parser.find_column(inventory_df, STORE_CANDIDATES)
        category_col = data_parser.find_column(inventory_df, CATEGORY_CANDIDATES)
        quantity_col = data_parser.find_column(inventory_df, AVAILABLE_QTY_CANDIDATES)
        amount_col = data_parser.find_column(inventory_df, AVAILABLE_AMOUNT_CANDIDATES)

        missing = {
            "仓库": warehouse_col,
            "门店": store_col,
            "一级分类": category_col,
            "可用数量": quantity_col,
            "可用金额": amount_col,
        }
        missing_fields = [name for name, col in missing.items() if col is None]
        if missing_fields:
            logger.error(f"缺少必要字段: {missing_fields}")
            return None

        # 剔除指定仓库
        filtered_df = inventory_df[~inventory_df[warehouse_col].isin(EXCLUDED_WAREHOUSES)].copy()
        removed_count = len(inventory_df) - len(filtered_df)
        logger.info(f"剔除仓库数据 {removed_count} 行")

        # 汇总透视（仅保留门店与分类）
        summary_df = (
            filtered_df
            .groupby([store_col, category_col], dropna=False)[[quantity_col, amount_col]]
            .sum(min_count=1)
            .reset_index()
        )

        summary_df = summary_df.rename(
            columns={
                store_col: "门店",
                category_col: "一级分类",
                quantity_col: "库存数量",
                amount_col: "库存金额",
            }
        )

        # 清理无效分类
        summary_df["一级分类"] = summary_df["一级分类"].astype(str).str.strip()
        summary_df = summary_df[summary_df["一级分类"] != ""]
        if summary_df.empty:
            logger.warning("过滤后无有效的门店分类数据")
            return None

        summary_df["库存数量"] = summary_df["库存数量"].fillna(0).astype(float)
        summary_df["库存金额"] = summary_df["库存金额"].fillna(0).astype(float)

        # 加工配送分析汇总并合入明细
        delivery_summary = build_delivery_summary(data_loader, data_parser)
        if delivery_summary is not None and not delivery_summary.empty:
            delivery_summary = delivery_summary.rename(columns={"调出门店": "门店"})
            summary_df = summary_df.merge(
                delivery_summary,
                on=["门店", "一级分类"],
                how="left",
            )
        else:
            summary_df["配送数量"] = 0
            summary_df["配送金额"] = 0

        summary_df["配送数量"] = summary_df.get("配送数量", 0).fillna(0).astype(float)
        summary_df["配送金额"] = summary_df.get("配送金额", 0).fillna(0).astype(float)

        summary_df["库存金额周转"] = np.where(
            summary_df["配送金额"] == 0,
            0,
            summary_df["库存金额"] / summary_df["配送金额"],
        )
        summary_df["库存数量周转"] = np.where(
            summary_df["配送数量"] == 0,
            0,
            summary_df["库存数量"] / summary_df["配送数量"],
        )

        # 构建分类和门店排序
        category_order_map = {
            cat: idx for idx, cat in enumerate(sorted(summary_df["一级分类"].unique()))
        }
        store_order_map = {
            store: idx for idx, store in enumerate(sorted(summary_df["门店"].unique()))
        }

        # 追加门店合计
        store_totals = (
            summary_df.groupby("门店")[
                ["库存数量", "库存金额", "配送数量", "配送金额"]
            ]
            .sum(min_count=1)
            .reset_index()
        )
        store_totals["一级分类"] = "合计"
        store_totals["store_order"] = store_totals["门店"].map(store_order_map)
        store_totals["segment_order"] = 1
        store_totals["category_order"] = len(category_order_map)
        store_totals["库存金额周转"] = np.where(
            store_totals["配送金额"] == 0,
            0,
            store_totals["库存金额"] / store_totals["配送金额"],
        )
        store_totals["库存数量周转"] = np.where(
            store_totals["配送数量"] == 0,
            0,
            store_totals["库存数量"] / store_totals["配送数量"],
        )

        # 总合计（按分类）
        category_totals = (
            summary_df.groupby("一级分类")[[
                "库存数量",
                "库存金额",
                "配送数量",
                "配送金额",
            ]]
            .sum(min_count=1)
            .reset_index()
        )
        category_totals["门店"] = "总合计"
        category_totals["store_order"] = len(store_order_map)
        category_totals["segment_order"] = 0
        category_totals["category_order"] = category_totals["一级分类"].map(category_order_map)
        category_totals["库存金额周转"] = np.where(
            category_totals["配送金额"] == 0,
            0,
            category_totals["库存金额"] / category_totals["配送金额"],
        )
        category_totals["库存数量周转"] = np.where(
            category_totals["配送数量"] == 0,
            0,
            category_totals["库存数量"] / category_totals["配送数量"],
        )

        grand_total = pd.DataFrame({
            "门店": ["总合计"],
            "一级分类": ["合计"],
            "库存数量": [category_totals["库存数量"].sum()],
            "库存金额": [category_totals["库存金额"].sum()],
            "配送数量": [category_totals["配送数量"].sum()],
            "配送金额": [category_totals["配送金额"].sum()],
            "store_order": [len(store_order_map)],
            "segment_order": [1],
            "category_order": [len(category_order_map)],
        })
        grand_total["库存金额周转"] = np.where(
            grand_total["配送金额"] == 0,
            0,
            grand_total["库存金额"] / grand_total["配送金额"],
        )
        grand_total["库存数量周转"] = np.where(
            grand_total["配送数量"] == 0,
            0,
            grand_total["库存数量"] / grand_total["配送数量"],
        )

        # 门店详细行
        detail_df = summary_df.copy()
        detail_df["store_order"] = detail_df["门店"].map(store_order_map)
        detail_df["segment_order"] = 0
        detail_df["category_order"] = detail_df["一级分类"].map(category_order_map)

        combined_df = pd.concat(
            [detail_df, store_totals, category_totals, grand_total],
            ignore_index=True,
        )

        combined_df = combined_df.sort_values(
            by=["store_order", "门店", "segment_order", "category_order", "一级分类"],
            kind="mergesort",
        ).reset_index(drop=True)

        pivot_df = combined_df[[
            "门店",
            "一级分类",
            "配送金额",
            "配送数量",
            "库存金额",
            "库存数量",
            "库存金额周转",
            "库存数量周转",
        ]]

        numeric_columns = [
            "配送金额",
            "配送数量",
            "库存金额",
            "库存数量",
            "库存金额周转",
            "库存数量周转",
        ]
        pivot_df[numeric_columns] = pivot_df[numeric_columns].apply(lambda col: col.round(1))

        report_date = pd.Timestamp.now().strftime("%Y-%m-%d")
        base_filename = f"订单库存{report_date}.xlsx"
        output_path = PROCESSED_DIR / base_filename
        output_path.parent.mkdir(parents=True, exist_ok=True)

        suffix = 1
        while output_path.exists():
            output_path = PROCESSED_DIR / f"订单库存{report_date}_{suffix}.xlsx"
            suffix += 1

        pivot_df.to_excel(output_path, index=False)

        apply_sheet_style(output_path)

        logger.info(f"库存门店分类透视报表生成成功: {output_path}")
        if delivery_summary is not None and not delivery_summary.empty:
            logger.info("配送分析分类汇总已合并并参与合计计算")
        print(f"[完成] 库存门店分类透视报表: {output_path.name}")
        return output_path

    except Exception as exc:
        logger.error(f"生成库存门店分类透视报表失败: {exc}")
        return None


def build_delivery_summary(data_loader, data_parser) -> Optional[pd.DataFrame]:
    """读取配送分析数据并按门店+一级分类汇总"""
    try:
        delivery_df = data_loader.load_latest_module_data("delivery_analysis")
        if delivery_df is None or delivery_df.empty:
            logger.warning("未获取到配送分析数据，跳过配送汇总")
            return None

        cleaned_df = clean_delivery_header(delivery_df)
        if cleaned_df is None or cleaned_df.empty:
            logger.warning("配送分析数据清洗后为空，跳过配送汇总")
            return None

        store_col = data_parser.find_column(cleaned_df, DELIVERY_STORE_CANDIDATES)
        category_col = data_parser.find_column(cleaned_df, DELIVERY_CATEGORY_CANDIDATES)
        qty_col = data_parser.find_column(cleaned_df, DELIVERY_QTY_CANDIDATES)
        amount_col = data_parser.find_column(cleaned_df, DELIVERY_AMOUNT_CANDIDATES)

        missing = {
            "调出门店": store_col,
            "一级分类": category_col,
            "配送数量": qty_col,
            "配送金额": amount_col,
        }
        missing_fields = [name for name, col in missing.items() if col is None]
        if missing_fields:
            logger.warning(f"配送分析缺少必要字段，跳过配送汇总: {missing_fields}")
            return None

        working_df = cleaned_df.dropna(subset=[store_col, category_col], how="any").copy()
        working_df[category_col] = working_df[category_col].astype(str).str.strip()
        working_df = working_df[working_df[category_col] != ""]
        if working_df.empty:
            logger.warning("配送分析有效分类数据为空，跳过配送汇总")
            return None

        for numeric_col in [qty_col, amount_col]:
            working_df[numeric_col] = (
                working_df[numeric_col]
                .astype(str)
                .str.replace(",", "", regex=False)
                .str.replace(" ", "", regex=False)
            )
            working_df[numeric_col] = pd.to_numeric(working_df[numeric_col], errors="coerce").fillna(0)

        summary_df = (
            working_df
            .groupby([store_col, category_col], dropna=False)[[qty_col, amount_col]]
            .sum(min_count=1)
            .reset_index()
        )

        summary_df = summary_df.rename(
            columns={
                store_col: "调出门店",
                category_col: "一级分类",
                qty_col: "配送数量",
                amount_col: "配送金额",
            }
        )

        summary_df = summary_df.sort_values(["调出门店", "一级分类"]).reset_index(drop=True)
        return summary_df

    except Exception as exc:
        logger.warning(f"配送分析汇总失败: {exc}")
        return None


def load_product_categories(data_loader) -> Optional[pd.DataFrame]:
    """加载商品分类数据"""
    try:
        category_df = data_loader.load_latest_module_data("product_archive")
        if category_df is None or category_df.empty:
            logger.warning("未获取到商品分类数据，后续分类列可能缺失")
            return None

        required_columns = ["商品代码", "一级分类"]
        missing_cols = [col for col in required_columns if col not in category_df.columns]
        if missing_cols:
            logger.warning(f"分类数据缺少字段: {missing_cols}")
            return None

        category_clean = category_df[required_columns].dropna(subset=["商品代码"])
        return category_clean.drop_duplicates(subset=["商品代码"])
    except Exception as exc:
        logger.warning(f"加载商品分类数据异常: {exc}")
        return None


def add_product_categories(inventory_df: pd.DataFrame, category_df: pd.DataFrame) -> pd.DataFrame:
    """为库存数据补充一级分类"""
    if category_df is None or category_df.empty:
        return inventory_df

    data_parser = get_data_parser()
    product_code_col = data_parser.find_column(inventory_df, PRODUCT_CODE_CANDIDATES)
    if product_code_col is None:
        logger.warning("库存数据缺少商品代码字段，无法补充分类")
        return inventory_df

    merged = inventory_df.merge(
        category_df[["商品代码", "一级分类"]],
        left_on=product_code_col,
        right_on="商品代码",
        how="left",
    )

    merged["一级分类"] = merged["一级分类"].fillna("")
    return merged.drop(columns=["商品代码"], errors="ignore")


def apply_sheet_style(file_path: Path) -> None:
    """应用统一的表格样式"""
    try:
        workbook = load_workbook(file_path)
        thin_side = Side(style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        font = Font(name="微软雅黑", size=10)

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.border = border

        workbook.save(file_path)
        workbook.close()
    except Exception as exc:
        logger.warning(f"应用样式失败: {exc}")


def clean_delivery_header(df: pd.DataFrame) -> pd.DataFrame:
    """去掉配送分析导出的标题行并恢复表头"""
    working = df.copy()
    working = working.dropna(how="all")
    if working.empty:
        return working

    header_row = working.iloc[0].fillna("")
    working = working.iloc[1:].reset_index(drop=True)
    working.columns = header_row
    working = working.dropna(how="all").reset_index(drop=True)
    return working
